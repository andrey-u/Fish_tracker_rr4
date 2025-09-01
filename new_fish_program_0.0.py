import pyautogui
import time
import tkinter as tk
import keyboard
import pytesseract
import cv2
import openpyxl
from PIL import Image
import pandas as pd
import os
import xlsxwriter
import datetime
import json
 

# основные
screenshot_interval = 0.2 # интервал между скриншотами
stop_flag=False
last_fish_name="" # символы, распознанные на последнем скриншоте
last_fish_weight=0
screen_width = 1920
screen_height = 1080

# для экселя
column_names = ["Название рыбы", "Масса (г)", "Номер удочки", "Дата поимки", "Цена"] # названия столбцов в файле эксель
row_to_insert = 2 # строчка, в которую будут добавляться новые рыбы
total_columns = len(column_names) # всего колонок
default_sheet_name = "Sheet1"
xlsx_filename = "Test_excel_file.xlsx" # имя эксель файла
xlsx_dirname = "excel_files" # имя директории (папки) для эксель файла
xlsx_path_to_file = os.path.join(xlsx_dirname, xlsx_filename) # путь к эксель файлу (относительный, относительно местоположения программы)

# для json
json_dirname = "fish_info"

name = "name"
json_uncounted_weight = "uncounted_weight"
json_uncounted_cost = "uncounted_cost"
uncounted_cost = 0.3
json_counted_weight = "counted_weight"
json_counted_cost = "counted_cost"
json_chat_weight = "chat_weight"
json_chat_cost = "chat_cost"
json_trof_weight = "trof_weight"
json_trof_cost = "trof_cost"
json_rare_trof_weight = "rare_trof_weight"
json_rare_trof_cost = "rare_trof_cost"
json_max_weight = "max_weight"


# другое
last_fishing_rod_number = 0


# список рыб
list_fish_names = [
                "Акула гигантская", 
                "Акула гренландская полярная", 
                "Акула плащеносная", 
                "Акула сельдевая атлантическая", 
                "Амур белый", 
                "Амур белый альбинос", 
                "Амур чёрный", 
                "Белоглазка", 
                "Белорыбица", 
                "Белуга каспийская", 
                "Белуга черноморская", 
                "Бельдюга европейская", 
                "Берикс красный",
                "Берш", 
                "Буффало большеротый", 
                "Буффало чёрный", 
                "Валёк", 
                "Вармоус", 
                "Вобла", 
                "Вырезуб", 
                "Вьюн", 
                "Гимантолоф атлантический", 
                "Голавль", 
                "Голец арктический", 
                "Голец Дрягина", 
                "Голец Куорский", 
                "Голец Леванидова", 
                "Голец сибирский-усач", 
                "Гольян", 
                "Гольян озёрный", 
                "Горбуша", 
                "Горбыль речной", 
                "Горбыль серебристый", 
                "Гребешок исландский", 
                "Густера", 
                "Доросома северная", 
                "Дрейссена речная", 
                "Елец", 
                "Елец сибирский", 
                "Ёрш", 
                "Ёрш-носарь", 
                "Жерех", 
                "Зубатка полосатая", 
                "Зубатка пятнистая", 
                "Зубатка синяя",
                "Калуга", 
                "Кальмар обыкновенный", 
                "Камбала морская", 
                "Камбала палтусовидная", 
                "Камбала хоботная", 
                "Карасекарп", 
                "Карась золотой", 
                "Карась серебряный", 
                "Карп голый", 
                "Карп голый - альбинос", 
                "Карп голый - призрак", 
                "Карп Динкенбюльский голый", 
                "Карп Динкенбюльский зеркальный", 
                "Карп Динкенбюльский линейный", 
                "Карп зеркальный", 
                "Карп зеркальный - альбинос", 
                "Карп зеркальный - призрак", 
                "Карп Кои Ёцусиро", 
                "Карп Кои Кохаку", 
                "Карп Кои Мамэсибори Госики", 
                "Карп Кои Мидори-гои", 
                "Карп Кои Наруми Асаги", 
                "Карп Кои Орэндзи Огон", 
                "Карп Кои Хи Уцури",                
                "Карп красный Старвас - зеркальный", 
                "Карп красный Старвас - чешуйчатый", 
                "Карп линейный", 
                "Карп линейный - альбинос", 
                "Карп линейный - призрак", 
                "Карп рамчатый", 
                "Карп рамчатый - альбинос", 
                "Карп рамчатый - призрак", 
                "Карп Супер Фрикс", 
                "Карп чешуйчатый", 
                "Карп чешуйчатый - альбинос", 
                "Карп чешуйчатый - призрак", 
                "Катран", 
                "Керчак европейский", 
                "Кета", 
                "Кижуч", 
                "Колюшка девятииглая", 
                "Колюшка малая южная", 
                "Колюшка трёхиглая", 
                "Конгер", 
                "Корюшка", 
                "Корюшка азиатская", 
                "Краб камчатский", 
                "Краб съедобный", 
                "Краппи белый", 
                "Краппи чёрный", 
                "Краснопёр монгольский", 
                "Краснопёр-Угай крупночешуйчатый", 
                "Краснопёрка", 
                "Кунджа", 
                "Кутум", 
                "Лаврак полосатый", 
                "Лаврак полосатый гибридный", 
                "Ленок острорылый", 
                "Ленок тупорылый", 
                "Лещ", 
                "Лещ восточный", 
                "Ликод полуголый", 
                "Ликод Эсмарка", 
                "Линь", 
                "Линь золотистый", 
                "Линь Квольсдорфский", 
                "Лосось атлантический", 
                "Лосось каспийский", 
                "Лосось ладожский", 
                "Лягушка", 
                "Макрель атлантическая", 
                "Макрурус северный", 
                "Мальма", 
                "Менёк", 
                "Мерланг", 
                "Мерлуза", 
                "Меч-рыба", 
                "Мидия", 
                "Микижа", 
                "Минога дальневосточная ручьевая", 
                "Минога каспийская", 
                "Минога сибирская", 
                "Минога трёхзубая", 
                "Минога украинская", 
                "Мольва голубая", 
                "Мольва обыкновенная", 
                "Морской чёрт", 
                "Муксун", 
                "Налим", 
                "Нейва", 
                "Нельма", 
                "Нерка", 
                "Окунь", 
                "Окунь белый", 
                "Окунь большеротый", 
                "Окунь каменный", 
                "Окунь малоротый", 
                "Окунь морской золотистый", 
                "Окунь морской норвежский", 
                "Окунь павлиний", 
                "Окунь пятнистый", 
                "Окунь солнечный", 
                "Окунь-клювач", 
                "Омуль арктический", 
                "Омуль байкальский", 
                "Опах краснопёрый", 
                "Осётр балтийский", 
                "Осётр восточносибирский", 
                "Осётр ладожский", 
                "Осётр персидский", 
                "Осётр русский", 
                "Палия кряжевая", 
                "Палия лудожная", 
                "Палия обыкновенная", 
                "Палтус атлантический", 
                "Палтус синекорый", 
                "Панцирник пятнистый", 
                "Пелядь", 
                "Перловица", 
                "Пескарь обыкновенный", 
                "Пескарь сибирский", 
                "Пикша", 
                "Пинагор", 
                "Плотва обыкновенная", 
                "Плотва сибирская", 
                "Подкаменщик сибирский", 
                "Подуст", 
                "Поллак", 
                "Пузанок каспийский", 
                "Путассу северная", 
                "Рак речной", 
                "Рипус", 
                "Ротан", 
                "Рыбец", 
                "Ряпушка", 
                "Ряпушка сибирская", 
                "Сазан", 
                "Сайда", 
                "Сайра атлантическая", 
                "Сардина европейская", 
                "Севрюга", 
                "Сельдь атлантическая", 
                "Сельдь Бражникова", 
                "Сельдь Кесслера", 
                "Сельдь черноморская", 
                "Сиг валаамский", 
                "Сиг волховский", 
                "Сиг вуоксинский", 
                "Сиг куорский", 
                "Сиг ладожский озёрный", 
                "Сиг острожский", 
                "Сиг свирский", 
                "Сиг чёрный", 
                "Сиг-лудога", 
                "Сиг-пыжьян", 
                "Сима", 
                "Сима жилая", 
                "Синец", 
                "Скат колючий", 
                "Скат полярный", 
                "Солнечник красноухий", 
                "Солнечник синежаберный", 
                "Сом", 
                "Сом альбинос", 
                "Сом амурский", 
                "Сомик канальный", 
                "Сомик оливковый", 
                "Стерлядь", 
                "Стерлядь сибирская", 
                "Судак", 
                "Судак светлопёрый", 
                "Таймень", 
                "Тарань", 
                "Толстолобик белый", 
                "Толстолобик пёстрый", 
                "Треска атлантическая", 
                "Тугун", 
                "Тунец голубой", 
                "Тюлька черноморская", 
                "Тюрбо", 
                "Угольщик обыкновенный", 
                "Угорь", 
                "Уклейка", 
                "Усач альбинос", 
                "Усач короткоголовый", 
                "Усач обыкновенный", 
                "Устрица съедобная", 
                "Форель озерная", 
                "Форель радужная", 
                "Форель ручьевая", 
                "Форель севанская", 
                "Хариус восточносибирский", 
                "Хариус европейский", 
                "Хариус западносибирский", 
                "Химера европейская", 
                "Центролоф чёрный", 
                "Чавыча", 
                "Чехонь", 
                "Чир", 
                "Чукучан", 
                "Шемая каспийская", 
                "Шемая черноморская", 
                "Шип", 
                "Щука обыкновенная", 
                "Щука панцирная", 
                "Язь"
            ]

# словарь рыб (для соотношения рыбы с файлом .json) (старый)
dict_fish_old={
    "Акула гигантская" : "Giant_shark",
    "Акула гренландская полярная" : "Greenland_shark",
    "Сом альбинос" : "Albino_catfish",
    "Усач альбинос" : "Albino_moustache",
    "Сом амурский" : "Amur_catfish",
    "Амур белый альбинос" : "Amur_white_albino",
    "Голец арктический" : "Arctic_char",
    "Омуль арктический" : "Arctic_omul",
    "Скат полярный" : "Arctic_stingray",
    "Корюшка азиатская" : "Asian_smelt",
    "Жерех" : "Asp",
    "Треска атлантическая" : "Atlantic_cod",
    "Палтус атлантический" : "Atlantic_halibut",
    "Сельдь атлантическая" : "Atlantic_herring",
    "Макрель атлантическая" : "Atlantic_mackerel",
    "Акула сельдевая атлантическая" : "Atlantic_porbeagle_shark",
    "Лосось атлантический" : "Atlantic_salmon",
    "Сайра атлантическая" : "Atlantic_saury",
    "Омуль байкальский" : "Baikal_omul",
    "Осётр балтийский" : "Baltic_sturgeon",
    "Берикс красный" : "Berix_red",
    "Берш" : "Bersh",
    "Буффало большеротый" : "Bigmouth_Buffalo",
    "Амур чёрный" : "Black_Amur",
    "Краппи чёрный" : "Black_crappie",
    "Белуга черноморская" : "Black_Sea_Beluga",
    "Сельдь черноморская" : "Black_Sea_Herring",
    "Тюлька черноморская" : "Black_Sea_sprat",
    "Сиг чёрный" : "Black_whitefish",
    "Уклейка" : "Bleak",
    "Палтус синекорый" : "Blue-skinned_halibut",
    "Тунец голубой" : "Bluefin_tuna",
    "Солнечник синежаберный" : "Bluegill_sunfish",
    "Синец" : "Blue_bream",
    "Мольва голубая" : "Blue_linnet",
    "Ленок тупорылый" : "Blunt-nosed_flax",
    "Сельдь Бражникова" : "Brazhnikov's_Herring",
    "Лещ" : "Bream",
    "Форель ручьевая" : "Brook_trout",
    "Буффало чёрный" : "Buffalo_black",
    "Налим" : "Burbot",
    "Сазан" : "Carp",
    "Карп Кои Хи Уцури" : "Carp_Koi_Hi_Utsuri",
    "Карп Кои Мидори-гои" : "Carp_Koi_Midori-goi",
    "Карп Кои Ёцусиро" : "Carp_Koi_Yotsushiro",
    "Карп линейный" : "Carp_linear",
    "Карп линейный - призрак" : "Carp_linear_-_ghost",
    "Карп чешуйчатый - призрак" : "Carp_scaly_-_ghost",
    "Карп Супер Фрикс" : "Carp_Super_Frix",
    "Белуга каспийская" : "Caspian_Beluga",
    "Минога каспийская" : "Caspian_lamprey",
    "Пузанок каспийский" : "Caspian_puzanok",
    "Лосось каспийский" : "Caspian_salmon",
    "Сом" : "Catfish",
    "Центролоф чёрный" : "Centrolophus_black",
    "Сомик канальный" : "Channel_catfish",
    "Чехонь" : "Chekhon",
    "Чавыча" : "Chinook_salmon",
    "Чир" : "Chir",
    "Голавль" : "Chub",
    "Чукучан" : "Chukuchan",
    "Кета" : "Chum_salmon",
    "Кижуч" : "Coho_salmon",
    "Усач обыкновенный" : "Common_barbel",
    "Палия обыкновенная" : "Common_char",
    "Угольщик обыкновенный" : "Common_coal_miner",
    "Пескарь обыкновенный" : "Common_gudgeon",
    "Мольва обыкновенная" : "Common_lingonberry",
    "Щука обыкновенная" : "Common_Pike",
    "Плотва обыкновенная" : "Common_roach",
    "Кальмар обыкновенный" : "Common_squid",
    "Конгер" : "Conger",
    "Елец" : "Dace",
    "Карп Динкенбюльский линейный" : "Dinkenbul_linear_carp",
    "Карп Динкенбюльский зеркальный" : "Dinkenbühl_Mirror_Carp",
    "Карп Динкенбюльский голый" : "Dinkenbühl_naked_carp",
    "Дрейссена речная" : "Dreissena_riverina",
    "Лещ восточный" : "Eastern_bream",
    "Хариус восточносибирский" : "East_Siberian_grayling",
    "Осётр восточносибирский" : "East_Siberian_sturgeon",
    "Краб съедобный" : "Edible_crab",
    "Устрица съедобная" : "Edible_oyster",
    "Угорь" : "Eel",
    "Ликод Эсмарка" : "Esmark's_Licode",
    "Химера европейская" : "European_chimera",
    "Бельдюга европейская" : "European_eelpout",
    "Хариус европейский" : "European_grayling",
    "Сардина европейская" : "European_sardine",
    "Керчак европейский" : "European_sculpin",
    "Минога дальневосточная ручьевая" : "Far_Eastern_brook_lamprey",
    "Карп рамчатый" : "Framed_carp",
    "Карп рамчатый - альбинос" : "Framed_carp_-_albino",
    "Карп рамчатый - призрак" : "Framed_carp_-_ghost",
    "Акула плащеносная" : "Frilled_shark",
    "Лягушка" : "Frog",
    "Карась золотой" : "Golden_crucian_carp",
    "Карась серебряный" : "Silver_crucian_carp",
    "Окунь морской золотистый" : "Golden_sea_bass",
    "Линь золотистый" : "Golden_Tench",
    "Голец Дрягина" : "Golets_Dryagina",
    "Голец Куорский" : "Golets_Kuorsky",
    "Голец Леванидова" : "Golets_Levanidova",
    "Пикша" : "Haddock",
    "Мерлуза" : "Hake",
    "Камбала палтусовидная" : "Halibut_flounder",
    "Гимантолоф атлантический" : "Hymantolophus_atlantica",
    "Гребешок исландский" : "Icelandic_scallop",
    "Язь" : "Ide",
    "Калуга" : "Kaluga",
    "Краб камчатский" : "Kamchatka_crab",
    "Карасекарп" : "Karasekarp",
    "Катран" : "Katran",
    "Сельдь Кесслера" : "Kessler's_Herring",
    "Карп Кои Мамэсибори Госики" : "Koi_Carp_Mameshibori_Goshiki",
    "Карп Кои Наруми Асаги" : "Koi_carp_Narumi_Asagi",
    "Карп Кои Орэндзи Огон" : "Koi_carp_Orenji_Ogon",
    "Карп Кои Кохаку" : "Koi_Kohaku_Carp",
    "Кунджа" : "Kunja",
    "Сиг куорский" : "Kuorsky_whitefish",
    "Кутум" : "Kutum",
    "Сиг ладожский озёрный" : "Ladoga_lake_whitefish",
    "Лосось ладожский" : "Ladoga_salmon",
    "Осётр ладожский" : "Ladoga_sturgeon",
    "Гольян озёрный" : "Lake_minnow",
    "Форель озерная" : "Lake_trout",
    "Окунь большеротый" : "Largemouth_bass",
    "Судак светлопёрый" : "Light-finned_pike-perch",
    "Ликод полуголый" : "Likod_half_naked",
    "Карп линейный - альбинос" : "Linear_carp_-_albino",
    "Линь Квольсдорфский" : "Lin_of_Kwolsdorf",
    "Вьюн" : "Loach",
    "Мальма" : "Malma",
    "Менёк" : "Menek",
    "Микижа" : "Mikizha",
    "Гольян" : "Minnow",
    "Карп зеркальный" : "Mirror_carp",
    "Карп зеркальный - альбинос" : "Mirror_carp_-_albino",
    "Карп зеркальный - призрак" : "Mirror_carp_-_ghost",
    "Краснопёр монгольский" : "Mongolian_rudd",
    "Морской чёрт" : "Monkfish",
    "Муксун" : "Muksun",
    "Мидия" : "Mussel",
    "Карп голый" : "Naked_carp",
    "Карп голый - альбинос" : "Naked_carp_-_albino",
    "Карп голый - призрак" : "Naked_ghost_Carp",
    "Нейва" : "Neiva",
    "Нельма" : "Nelma",
    "Колюшка девятииглая" : "Nine-spined_stickleback",
    "Путассу северная" : "Northern_blue_whiting",
    "Доросома северная" : "Northern_dorosoma",
    "Макрурус северный" : "Northern_Macrurus",
    "Окунь морской норвежский" : "Norwegian_sea_bass",
    "Сомик оливковый" : "Olive_catfish",
    "Опах краснопёрый" : "Opah_redfin",
    "Сиг острожский" : "Ostrog_whitefish",
    "Палия лудожная" : "Palia_ludozhnaya",
    "Окунь павлиний" : "Peacock_bass",
    "Перловица" : "Pearl_barley",
    "Пелядь" : "Peled",
    "Окунь-клювач" : "Perch-beaker",
    "Окунь" : "Perch",
    "Осётр персидский" : "Persian_sturgeon",
    "Щука панцирная" : "Pike_armored",
    "Пинагор" : "Pinagor",
    "Горбуша" : "Pink_salmon",
    "Палия кряжевая" : "Platy_char",
    "Подуст" : "Podust",
    "Поллак" : "Pollack",
    "Камбала хоботная" : "Proboscis_flounder",
    "Валёк" : "Prosopium",
    "Форель радужная" : "Rainbow_trout",
    "Солнечник красноухий" : "Red-eared_sunfish",
    "Краснопёр-Угай крупночешуйчатый" : "Redfin-Eugay_large-scaled",
    "Карп красный Старвас - зеркальный" : "Red_carp_Starvas_-_mirror",
    "Карп красный Старвас - чешуйчатый" : "Red_carp_Starvas_-_scaly",
    "Нерка" : "Red_salmon",
    "Рипус" : "Ripus",
    "Рак речной" : "River_crayfish",
    "Горбыль речной" : "River_slab",
    "Окунь каменный" : "Rock_perch",
    "Ротан" : "Rotan",
    "Краснопёрка" : "Rudd",
    "Ёрш-носарь" : "Ruff-nose",
    "Ёрш" : "Ruff",
    "Осётр русский" : "Russian_sturgeon",
    "Рыбец" : "Rybets",
    "Сайда" : "Saida",
    "Карп чешуйчатый" : "Scaly_carp",
    "Карп чешуйчатый - альбинос" : "Scaly_carp_-_albino",
    "Камбала морская" : "Sea_flounder",
    "Форель севанская" : "Sevan_trout",
    "Ленок острорылый" : "Sharp-nosed_flax",
    "Шемая черноморская" : "Shemaya_Black_Sea",
    "Шемая каспийская" : "Shemaya_Caspian",
    "Шип" : "Ship",
    "Усач короткоголовый" : "Short-headed_barbel",
    "Голец сибирский-усач" : "Siberian_barbel",
    "Елец сибирский" : "Siberian_dace",
    "Пескарь сибирский" : "Siberian_gudgeon",
    "Минога сибирская" : "Siberian_lamprey",
    "Плотва сибирская" : "Siberian_roach",
    "Подкаменщик сибирский" : "Siberian_sculpin",
    "Стерлядь сибирская" : "Siberian_sterlet",
    "Ряпушка сибирская" : "Siberian_vendace",
    "Сиг-лудога" : "Sig-ludoga",
    "Сиг-пыжьян" : "Sig-pyzhyan",
    "Толстолобик пёстрый" : "Silver_carp",
    "Толстолобик белый" : "Silver_carp_white",
    "Горбыль серебристый" : "Silver_slab",
    "Сима" : "Sima",
    "Сима жилая" : "Sima_zhilaya",
    "Окунь малоротый" : "Smallmouth_bass",
    "Корюшка" : "Smelt",
    "Колюшка малая южная" : "Southern_small_stickleback",
    "Скат колючий" : "Spiny_stingray",
    "Окунь пятнистый" : "Spotted_bass",
    "Зубатка пятнистая" : "Spotted_catfish",
    "Панцирник пятнистый" : "Spotted_Corymbose",
    "Севрюга" : "Stellate_sturgeon",
    "Стерлядь" : "Sterlet",
    "Лаврак полосатый гибридный" : "Striped_bass_hybrid",
    "Зубатка полосатая" : "Striped_catfish",
    "Лаврак полосатый" : "Striped_sea_bass",
    "Окунь солнечный" : "Sunfish",
    "Сиг свирский" : "Svirsky_whitefish",
    "Меч-рыба" : "Swordfish",
    "Таймень" : "Taimen",
    "Тарань" : "Taran",
    "Линь" : "Tench",
    "Зубатка синяя" : "The_catfish_is_blue",
    "Вырезуб" : "The_notch",
    "Колюшка трёхиглая" : "Three-spined_stickleback",
    "Минога трёхзубая" : "Three-toothed_lamprey",
    "Тугун" : "Tugun",
    "Тюрбо" : "Turbot",
    "Минога украинская" : "Ukrainian_lamprey",
    "Сиг валаамский" : "Valaam_whitefish",
    "Ряпушка" : "Vendace",
    "Вобла" : "vobla",
    "Сиг волховский" : "Volkhov_whitefish",
    "Сиг вуоксинский" : "Vuoksi_whitefish",
    "Вармоус" : "Warmous",
    "Хариус западносибирский" : "West_Siberian_grayling",
    "Белоглазка" : "White-eye",
    "Белорыбица" : "Whitefish",
    "Амур белый" : "White_Amur",
    "Густера" : "white_bream",
    "Краппи белый" : "White_crappie",
    "Окунь белый" : "White_perch",
    "Мерланг" : "Whiting",
    "Судак" : "Zander"
}

# словарь рыб (для соотношения рыбы с файлом .json)
dict_fish = {
                "Акула гигантская": "Giant_shark",
                "Акула гренландская полярная": "Greenland_shark",
                "Акула плащеносная": "Frilled_shark",
                "Акула сельдевая атлантическая": "Atlantic_porbeagle_shark",
                "Амур белый": "White_Amur",
                "Амур белый альбинос": "Amur_white_albino",
                "Амур чёрный": "Black_Amur",
                "Белоглазка": "White-eye",
                "Белорыбица": "Whitefish",
                "Белуга каспийская": "Caspian_Beluga",
                "Белуга черноморская": "Black_Sea_Beluga",
                "Бельдюга европейская": "European_eelpout",
                "Берикс красный": "Berix_red",
                "Берш": "Bersh",
                "Буффало большеротый": "Bigmouth_Buffalo",
                "Буффало чёрный": "Buffalo_black",
                "Валёк": "Prosopium",
                "Вармоус": "Warmous",
                "Вобла": "vobla",
                "Вырезуб": "The_notch",
                "Вьюн": "Loach",
                "Гимантолоф атлантический": "Hymantolophus_atlantica",
                "Голавль": "Chub",
                "Голец арктический": "Arctic_char",
                "Голец Дрягина": "Golets_Dryagina",
                "Голец Куорский": "Golets_Kuorsky",
                "Голец Леванидова": "Golets_Levanidova",
                "Голец сибирский-усач": "Siberian_barbel",
                "Гольян": "Minnow",
                "Гольян озёрный": "Lake_minnow",
                "Горбуша": "Pink_salmon",
                "Горбыль речной": "River_slab",
                "Горбыль серебристый": "Silver_slab",
                "Гребешок исландский": "Icelandic_scallop",
                "Густера": "white_bream",
                "Доросома северная": "Northern_dorosoma",
                "Дрейссена речная": "Dreissena_riverina",
                "Елец": "Dace",
                "Елец сибирский": "Siberian_dace",
                "Ёрш": "Ruff",
                "Ёрш-носарь": "Ruff-nose",
                "Жерех": "Asp",
                "Зубатка полосатая": "Striped_catfish",
                "Зубатка пятнистая": "Spotted_catfish",
                "Зубатка синяя": "The_catfish_is_blue",
                "Калуга": "Kaluga",
                "Кальмар обыкновенный": "Common_squid",
                "Камбала морская": "Sea_flounder",
                "Камбала палтусовидная": "Halibut_flounder",
                "Камбала хоботная": "Proboscis_flounder",
                "Карасекарп": "Karasekarp",
                "Карась золотой": "Golden_crucian_carp",
                "Карась серебряный": "Silver_crucian_carp",
                "Карп голый": "Naked_carp",
                "Карп голый - альбинос": "Naked_carp_-_albino",
                "Карп голый - призрак": "Naked_ghost_Carp",
                "Карп Динкенбюльский голый": "Dinkenbühl_naked_carp",
                "Карп Динкенбюльский зеркальный": "Dinkenbühl_Mirror_Carp",
                "Карп Динкенбюльский линейный": "Dinkenbul_linear_carp",
                "Карп зеркальный": "Mirror_carp",
                "Карп зеркальный - альбинос": "Mirror_carp_-_albino",
                "Карп зеркальный - призрак": "Mirror_carp_-_ghost",
                "Карп Кои Ёцусиро": "Carp_Koi_Yotsushiro",
                "Карп Кои Кохаку": "Koi_Kohaku_Carp",
                "Карп Кои Мамэсибори Госики": "Koi_Carp_Mameshibori_Goshiki",
                "Карп Кои Мидори-гои": "Carp_Koi_Midori-goi",
                "Карп Кои Наруми Асаги": "Koi_carp_Narumi_Asagi",
                "Карп Кои Орэндзи Огон": "Koi_carp_Orenji_Ogon",
                "Карп Кои Хи Уцури": "Carp_Koi_Hi_Utsuri",
                "Карп красный Старвас - зеркальный": "Red_carp_Starvas_-_mirror",
                "Карп красный Старвас - чешуйчатый": "Red_carp_Starvas_-_scaly",
                "Карп линейный": "Carp_linear",
                "Карп линейный - альбинос": "Linear_carp_-_albino",
                "Карп линейный - призрак": "Carp_linear_-_ghost",
                "Карп рамчатый": "Framed_carp",
                "Карп рамчатый - альбинос": "Framed_carp_-_albino",
                "Карп рамчатый - призрак": "Framed_carp_-_ghost",
                "Карп Супер Фрикс": "Carp_Super_Frix",
                "Карп чешуйчатый": "Scaly_carp",
                "Карп чешуйчатый - альбинос": "Scaly_carp_-_albino",
                "Карп чешуйчатый - призрак": "Carp_scaly_-_ghost",
                "Катран": "Katran",
                "Керчак европейский": "European_sculpin",
                "Кета": "Chum_salmon",
                "Кижуч": "Coho_salmon",
                "Колюшка девятииглая": "Nine-spined_stickleback",
                "Колюшка малая южная": "Southern_small_stickleback",
                "Колюшка трёхиглая": "Three-spined_stickleback",
                "Конгер": "Conger",
                "Корюшка": "Smelt",
                "Корюшка азиатская": "Asian_smelt",
                "Краб камчатский": "Kamchatka_crab",
                "Краб съедобный": "Edible_crab",
                "Краппи белый": "White_crappie",
                "Краппи чёрный": "Black_crappie",
                "Краснопёр монгольский": "Mongolian_rudd",
                "Краснопёр-Угай крупночешуйчатый": "Redfin-Eugay_large-scaled",
                "Краснопёрка": "Rudd",
                "Кунджа": "Kunja",
                "Кутум": "Kutum",
                "Лаврак полосатый": "Striped_sea_bass",
                "Лаврак полосатый гибридный": "Striped_bass_hybrid",
                "Ленок острорылый": "Sharp-nosed_flax", 
                "Ленок тупорылый": "Blunt-nosed_flax",
                "Лещ": "Bream",
                "Лещ восточный": "Eastern_bream",
                "Ликод полуголый": "Likod_half_naked",
                "Ликод Эсмарка": "Esmark's_Licode",
                "Линь": "Tench",
                "Линь золотистый": "Golden_Tench",
                "Линь Квольсдорфский": "Lin_of_Kwolsdorf",
                "Лосось атлантический": "Atlantic_salmon",
                "Лосось каспийский": "Caspian_salmon",
                "Лосось ладожский": "Ladoga_salmon",
                "Лягушка": "Frog",
                "Макрель атлантическая": "Atlantic_mackerel",
                "Макрурус северный": "Northern_Macrurus",
                "Мальма": "Malma",
                "Менёк": "Menek",
                "Мерланг": "Whiting",
                "Мерлуза": "Hake",
                "Меч-рыба": "Swordfish",
                "Мидия": "Mussel",
                "Микижа": "Mikizha",
                "Минога дальневосточная ручьевая": "Far_Eastern_brook_lamprey",
                "Минога каспийская": "Caspian_lamprey",
                "Минога сибирская": "Siberian_lamprey",
                "Минога трёхзубая": "Three-toothed_lamprey",
                "Минога украинская": "Ukrainian_lamprey",
                "Мольва голубая": "Blue_linnet",
                "Мольва обыкновенная": "Common_lingonberry",
                "Морской чёрт": "Monkfish",
                "Муксун": "Muksun",
                "Налим": "Burbot",
                "Нейва": "Neiva",
                "Нельма": "Nelma",
                "Нерка": "Red_salmon",
                "Окунь": "Perch",
                "Окунь белый": "White_perch",
                "Окунь большеротый": "Largemouth_bass",
                "Окунь каменный": "Rock_perch",
                "Окунь малоротый": "Smallmouth_bass",
                "Окунь морской золотистый": "Golden_sea_bass",
                "Окунь морской норвежский": "Norwegian_sea_bass",
                "Окунь павлиний": "Peacock_bass",
                "Окунь пятнистый": "Spotted_bass",
                "Окунь солнечный": "Sunfish",
                "Окунь-клювач": "Perch-beaker",
                "Омуль арктический": "Arctic_omul",
                "Омуль байкальский": "Baikal_omul",
                "Опах краснопёрый": "Opah_redfin",
                "Осётр балтийский": "Baltic_sturgeon",
                "Осётр восточносибирский": "East_Siberian_sturgeon",
                "Осётр ладожский": "Ladoga_sturgeon",
                "Осётр персидский": "Persian_sturgeon",
                "Осётр русский": "Russian_sturgeon",
                "Палия кряжевая": "Platy_char",
                "Палия лудожная": "Palia_ludozhnaya",
                "Палия обыкновенная": "Common_char",
                "Палтус атлантический": "Atlantic_halibut",
                "Палтус синекорый": "Blue-skinned_halibut",
                "Панцирник пятнистый": "Spotted_Corymbose",
                "Пелядь": "Peled",
                "Перловица": "Pearl_barley",
                "Пескарь обыкновенный": "Common_gudgeon",
                "Пескарь сибирский": "Siberian_gudgeon",
                "Пикша": "Haddock",
                "Пинагор": "Pinagor",
                "Плотва обыкновенная": "Common_roach",
                "Плотва сибирская": "Siberian_roach",
                "Подкаменщик сибирский": "Siberian_sculpin",
                "Подуст": "Podust",
                "Поллак": "Pollack",
                "Пузанок каспийский": "Caspian_puzanok",
                "Путассу северная": "Northern_blue_whiting",
                "Рак речной": "River_crayfish",
                "Рипус": "Ripus",
                "Ротан": "Rotan",
                "Рыбец": "Rybets",
                "Ряпушка": "Vendace",
                "Ряпушка сибирская": "Siberian_vendace",
                "Сазан": "Wild_carp",
                "Сайда": "Saida",
                "Сайра атлантическая": "Atlantic_saury",
                "Сардина европейская": "European_sardine",
                "Севрюга": "Stellate_sturgeon", # 3.313 стоит 11.50
                "Сельдь атлантическая": "Atlantic_herring",
                "Сельдь Бражникова": "Brazhnikov's_Herring",
                "Сельдь Кесслера": "Kessler's_Herring",
                "Сельдь черноморская": "Black_Sea_Herring",
                "Сиг валаамский": "Valaam_whitefish",
                "Сиг волховский": "Volkhov_whitefish",
                "Сиг вуоксинский": "Vuoksi_whitefish",
                "Сиг куорский": "Kuorsky_whitefish",
                "Сиг ладожский озёрный": "Ladoga_lake_whitefish",
                "Сиг острожский": "Ostrog_whitefish",
                "Сиг свирский": "Svirsky_whitefish",
                "Сиг чёрный": "Black_whitefish",
                "Сиг-лудога": "Sig-ludoga",
                "Сиг-пыжьян": "Sig-pyzhyan",
                "Сима": "Sima",
                "Сима жилая": "Sima_zhilaya",
                "Синец": "Blue_bream",
                "Скат колючий": "Spiny_stingray",
                "Скат полярный": "Arctic_stingray",
                "Солнечник красноухий": "Red-eared_sunfish",
                "Солнечник синежаберный": "Bluegill_sunfish",
                "Сом": "Catfish",
                "Сом альбинос": "Albino_catfish",
                "Сом амурский": "Amur_catfish",
                "Сомик канальный": "Channel_catfish",
                "Сомик оливковый": "Olive_catfish",
                "Стерлядь": "Sterlet",
                "Стерлядь сибирская": "Siberian_sterlet",
                "Судак": "Zander",
                "Судак светлопёрый": "Light-finned_pike-perch",
                "Таймень": "Taimen",
                "Тарань": "Taran",
                "Толстолобик белый": "Silver_carp_white",
                "Толстолобик пёстрый": "Silver_carp",
                "Треска атлантическая": "Atlantic_cod",
                "Тугун": "Tugun",
                "Тунец голубой": "Bluefin_tuna",
                "Тюлька черноморская": "Black_Sea_sprat",
                "Тюрбо": "Turbot",
                "Угольщик обыкновенный": "Common_coal_miner",
                "Угорь": "Eel",
                "Уклейка": "Bleak",
                "Усач альбинос": "Albino_moustache",
                "Усач короткоголовый": "Short-headed_barbel",
                "Усач обыкновенный": "Common_barbel",
                "Устрица съедобная": "Edible_oyster",
                "Форель озерная": "Lake_trout",
                "Форель радужная": "Rainbow_trout",
                "Форель ручьевая": "Brook_trout",
                "Форель севанская": "Sevan_trout",
                "Хариус восточносибирский": "East_Siberian_grayling",
                "Хариус европейский": "European_grayling",
                "Хариус западносибирский": "West_Siberian_grayling",
                "Химера европейская": "European_chimera",
                "Центролоф чёрный": "Centrolophus_black",
                "Чавыча": "Chinook_salmon",
                "Чехонь": "Chekhon",
                "Чир": "Chir",
                "Чукучан": "Chukuchan",
                "Шемая каспийская": "Shemaya_Caspian",
                "Шемая черноморская": "Shemaya_Black_Sea",
                "Шип": "Ship",
                "Щука обыкновенная": "Common_Pike",
                "Щука панцирная": "Pike_armored",
                "Язь": "Ide"
}


def count_fish_cost(fish_name: str, fish_weight: int):
    fish_weight = float(fish_weight/1000)
    path_to_fish_file = os.path.join(json_dirname, dict_fish[fish_name]+".json")
    with open (path_to_fish_file) as fish_file:
        fish_data_json = json.load(fish_file)
    
    curent_fish_name = fish_data_json[name]
    curent_fish_uncounted_weight = fish_data_json[json_uncounted_weight]
    curent_fish_uncounted_cost = fish_data_json[json_uncounted_cost]
    curent_fish_counted_weight = fish_data_json[json_counted_weight]
    curent_fish_counted_cost = fish_data_json[json_counted_cost]
    curent_fish_chat_weight = fish_data_json[json_chat_weight]
    curent_fish_chat_cost = fish_data_json[json_chat_cost]
    curent_fish_trof_weight = fish_data_json[json_trof_weight]
    curent_fish_trof_cost = fish_data_json[json_trof_cost]
    curent_fish_rare_trof_weight = fish_data_json[json_rare_trof_weight]
    curent_fish_rare_trof_cost = fish_data_json[json_rare_trof_cost]
    curent_fish_max_weight = fish_data_json[json_max_weight]

    if (fish_weight > curent_fish_max_weight or fish_weight == 0): return 0

    if (fish_weight >= curent_fish_rare_trof_weight and curent_fish_rare_trof_weight != 0):
        print("Редкий трофей!!!")
        total_cost = fish_weight * curent_fish_rare_trof_cost
        return total_cost

    elif (fish_weight >= curent_fish_trof_weight and curent_fish_trof_weight != 0):
        print("Трофей!")
        total_cost = fish_weight * curent_fish_trof_cost
        return total_cost
    
    elif (fish_weight >= curent_fish_chat_weight and curent_fish_chat_weight != 0):
        print("Чатная рыба")
        total_cost = fish_weight * curent_fish_chat_cost
        return total_cost
    
    elif (fish_weight >= curent_fish_counted_weight and curent_fish_counted_weight != 0):
        print("Зачетная рыба")
        total_cost = fish_weight * curent_fish_counted_cost
        return total_cost
    
    else :
        print("Незачетная рыба (или недостаточно информации)")
        total_cost = fish_weight * uncounted_cost
        tmp_int_cost = int(total_cost*100)
        total_cost = tmp_int_cost/100
        return total_cost

def set_last_fishing_rod_number(num): 
    global last_fishing_rod_number
    last_fishing_rod_number = num

def add_to_xlsx_file(list_of_data: list, filepath: str):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    ws.insert_rows(row_to_insert)
    col_to_insert = 1 # начинаем вставку данных с первой строки
    while (col_to_insert <= total_columns):
        ws.cell(row=row_to_insert, column=col_to_insert).value = list_of_data[col_to_insert-1]
        col_to_insert+=1
    wb.save(filepath)

def create_empty_excel(columns: list, filepath: str, sheet_name: str = default_sheet_name):
    df = pd.DataFrame(columns=columns)
    excel_writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(excel_writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 0))
    excel_writer._save()

    return filepath

def parse_excel_to_dict_list(filepath: str, sheet_name: str = default_sheet_name):
    # Загружаем Excel файл в DataFrame
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Преобразуем DataFrame в список словарей
    dict_list = df.to_dict(orient='records')

    return dict_list

def check_needed_files():
    if (not os.path.exists(xlsx_dirname)): 
        os.makedirs(xlsx_dirname)
        create_empty_excel(column_names, xlsx_path_to_file)
    elif (not os.path.exists(xlsx_path_to_file)):
        create_empty_excel(column_names, xlsx_path_to_file)


def image_processing():
    image = cv2.imread("screenshot.png")
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    thresh = cv2.threshold(gray, 215, 255, cv2.THRESH_BINARY_INV)[1]
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    result = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel, iterations=1)
    invert = cv2.bitwise_not(result)
    # cv2.imshow('invert', invert)
    # cv2.waitKey()
    # cv2.destroyAllWindows()
    return invert

def make_screenshot():
    screenshot = pyautogui.screenshot(region=(500,70,900,100))
    screenshot.save("screenshot.png")

def get_fish_data_from_screenshot():
    image = image_processing()
    fish_data = pytesseract.image_to_string(image, lang='rus', config='--psm 6').replace(",", ".")
    return fish_data

def start_making_screenshots():
    global last_fish_name
    global last_fish_weight
    while (not stop_flag):
        time.sleep(screenshot_interval)
        # Выход, если в это время был произведен выход
        if (stop_flag): break

        make_screenshot()
        fish_data = get_fish_data_from_screenshot()
        fish_name = fish_data.split("\n")[0] 
        fish_weight = 0
        
        # Проверяем, что на прошлом скриншоте была другая рыба, если не другая-выходим
        if last_fish_name == fish_name: continue
        # Обновляем последнюю пойманную рыбу
        last_fish_name = fish_name

        if fish_name in list_fish_names:
            # ската не нужно вращать
            if (fish_name != "Cкат полярный"):
                posx, posy = pyautogui.position()
                pyautogui.moveTo(screen_width/3.5, screen_height/2, duration=0.01)
                pyautogui.mouseDown()
                pyautogui.moveTo(screen_width/3.5, screen_height/2+100, duration=0.01)
                make_screenshot()  
                pyautogui.moveTo(screen_width/3.5, screen_height/2, duration=0.01)
                pyautogui.mouseUp()
                pyautogui.moveTo(posx,posy, duration=0.01) 
            else : make_screenshot()
            fish_data = get_fish_data_from_screenshot()
            try:
                fish_weight = str(fish_data.split("\n")[1].split("/")[0])
                # перевод из килограмм в граммы
                if ('к' in fish_weight):
                    fish_weight = fish_weight.replace("к", "").replace("г", "").replace(" ", "")
                    fish_weight = int(float(fish_weight)*1000)
                # если рыба уже в граммах
                elif ('г' in fish_weight):
                    fish_weight = fish_weight.replace("г", "").replace(" ", "")
                    fish_weight = int(fish_weight)
                else: fish_weight = 0 
            except:
                fish_weight = 0
            
            # устранение ошибок, связанных с нераспознованием запятой (может не сработать разве что на гигантскую акулу)
            json_path = str(os.path.join(json_dirname, dict_fish[fish_name])+".json")
            print(json_path)
            with open(json_path, "r") as json_file:
                json_fish_data = json.load(json_file)
                if (int(fish_weight) > int(json_fish_data[json_max_weight]*1000)): fish_weight = int(fish_weight)/1000
            
            # Дополнительно проверяем на ту же самую рыбу (если пойманы 2 рыбы одинаковой массы-сожалею)
            if (not last_fish_weight == fish_weight):
                print(fish_name)
                print(fish_weight)
                last_fish_weight = fish_weight

                fish_cost = float(count_fish_cost(fish_name=fish_name, fish_weight=fish_weight))
                fish_cost = round(fish_cost, 2)
                print("Цена: "+str(fish_cost))

                add_to_xlsx_file([fish_name, fish_weight, last_fishing_rod_number, datetime.datetime.now(), str(fish_cost).replace(".",",")], xlsx_path_to_file)
                
    

def set_stop_flag():
    global stop_flag
    print(datetime.datetime.now())
    stop_flag = True


def main():
    # проверка и создание необходимых эксель файлов
    check_needed_files()  
    # горячие клавиши для установки номера последней взятой удочки (можно поменять номера или добавить дополнительные)
    keyboard.add_hotkey('1', lambda: set_last_fishing_rod_number(1))
    keyboard.add_hotkey('2', lambda: set_last_fishing_rod_number(2))
    keyboard.add_hotkey('3', lambda: set_last_fishing_rod_number(3))

    # установка стоп-кнопки (временно)
    keyboard.add_hotkey('8', set_stop_flag)    

    print("программа работает")  
    # пауза между действиями мышки (для вращения рыбы)  
    pyautogui.PAUSE=0.03
    # установка размеров экрана
    global screen_width
    global screen_height
    screen_width, screen_height = pyautogui.size()
    # запуск основного цикла
    start_making_screenshots()

    

if __name__ == "__main__":
    main()



def create_excel_from_dict_list(dict_list: list, output_filename: str, sheet_name='Sheet1'):
    # Создаем директорию, если она не существует
    if not os.path.exists('excel_files'):
        os.makedirs('excel_files')

    filepath = os.path.join('excel_files', output_filename)

    # Создаем новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Записываем данные из списка словарей в Excel
    if dict_list:
        header = list(dict_list[0].keys())
        ws.append(header)  # Записываем заголовки

        for row in dict_list:
            ws.append([row[col] for col in header])

    

    # Автоматическое изменение ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Сохраняем файл
    wb.save(filepath)
    return filepath

def work_with_excel():
    pass
    # записывает данные в файл эксель
    # add_to_xlsx_file(["my fish",50,last_fishing_rod_number, datetime.datetime.now()], path_to_xlsx_file)
    # получает и выводит данные из файла эксель
    # dict_of_xlsx_file = parse_excel_to_dict_list(path_to_xlsx_file)
    # for elem in dict_of_xlsx_file: print(elem)
